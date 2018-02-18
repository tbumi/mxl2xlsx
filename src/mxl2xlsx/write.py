from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

COLUMN_WIDTH_IN_EXCEL = 4


def write_to_excel(file_path, excel_grid, title):
    workbook = Workbook()
    worksheet = workbook.active

    partitur_font = Font(name='Partitur', size=12)
    normal_text_font = Font(name='Calibri', size=11, bold=True)
    title_text_font = Font(name='Cambria', size=20, bold=True)
    center_alignment = Alignment(horizontal='center')
    left_alignment = Alignment(horizontal='left')
    left_border = Border(left=Side(style='thin'))
    right_border = Border(right=Side(style='thin'))

    # write the title
    worksheet.merge_cells(
        start_row=1, start_column=1, end_row=1,
        end_column=max(len(line) for line in excel_grid))
    tc = worksheet.cell(row=1, column=1, value=title)
    tc.font = title_text_font
    tc.alignment = center_alignment

    for row_num, row_entries in enumerate(excel_grid):
        for col_num, col_entries in enumerate(row_entries):
            worksheet.column_dimensions[get_column_letter(col_num+1)].width = COLUMN_WIDTH_IN_EXCEL
            c = worksheet.cell(
                row=row_num+1, column=col_num+1, value=col_entries['text'])
            if col_entries['format'] == 'partitur':
                c.font = partitur_font
                c.alignment = center_alignment
            elif col_entries['format'] == 'partitur_lborder':
                c.font = partitur_font
                c.alignment = center_alignment
                c.border = left_border
            elif col_entries['format'] == 'partitur_rborder':
                c.font = partitur_font
                c.alignment = center_alignment
                c.border = right_border
            else:
                c.font = normal_text_font
                c.alignment = left_alignment

    workbook.save(file_path)
